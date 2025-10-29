from __future__ import annotations

import asyncio
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import Branch, VehicleStock


EXCEL_HEADERS = [
    "branch_code",
    "branch_name",
    "city",
    "latitude",
    "longitude",
    "model_code",
    "model_name",
    "variant",
    "color",
    "quantity",
    "reserved",
]


@dataclass
class ImportSummary:
    processed: int = 0
    created: int = 0
    updated: int = 0
    removed: int = 0


class ExcelSyncService:
    """Synchronise vehicle stock with the canonical Excel workbook."""

    def __init__(self, session: AsyncSession, workbook_path: Path) -> None:
        self.session = session
        self.workbook_path = workbook_path

    async def import_inventory(self) -> ImportSummary:
        """Load or refresh vehicle stock entries from the Excel sheet."""
        workbook = await self._load_workbook()
        worksheet = workbook.active

        header = [cell for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        normalized = [self._normalize(value) for value in header]

        summary = ImportSummary()
        seen_rows: set[int] = set()

        for offset, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not any(row):
                continue

            summary.processed += 1
            row_data = {normalized[idx]: row[idx] for idx in range(len(normalized)) if idx < len(row)}
            excel_row_number = offset
            seen_rows.add(excel_row_number)

            stock = await self._get_stock_by_row(excel_row_number)

            payload = self._payload_from_row(row_data, excel_row_number)
            branch = await self._ensure_branch(payload)

            if stock is None:
                stock = VehicleStock(**payload)
                self.session.add(stock)
                summary.created += 1
            else:
                for field, value in payload.items():
                    setattr(stock, field, value)
                summary.updated += 1

            if branch and stock:
                stock.branch_code = branch.code
                stock.branch_name = branch.name
                stock.city = branch.city
                stock.latitude = branch.latitude
                stock.longitude = branch.longitude

        await self.session.flush()

        # Remove stale stock entries that no longer exist in the workbook.
        if seen_rows:
            stmt = delete(VehicleStock).where(VehicleStock.excel_row_number.notin(seen_rows))
            result = await self.session.execute(stmt)
            summary.removed = result.rowcount or 0

        await self.session.commit()
        await asyncio.to_thread(workbook.close)
        return summary

    async def push_stock_update(self, stock: VehicleStock) -> None:
        """Write the latest quantity/reserved values back to the workbook."""
        if stock.excel_row_number is None:
            return

        workbook = await self._load_workbook()
        worksheet = workbook.active

        row_index = stock.excel_row_number
        worksheet.cell(row=row_index, column=EXCEL_HEADERS.index("quantity") + 1, value=int(stock.quantity))
        worksheet.cell(row=row_index, column=EXCEL_HEADERS.index("reserved") + 1, value=int(stock.reserved))
        worksheet.cell(row=row_index, column=EXCEL_HEADERS.index("branch_name") + 1, value=stock.branch_name)
        worksheet.cell(row=row_index, column=EXCEL_HEADERS.index("city") + 1, value=stock.city)

        await asyncio.to_thread(workbook.save, self.workbook_path)
        await asyncio.to_thread(workbook.close)

    async def export_snapshot(self, target_path: Path | None = None) -> Path:
        """Generate a fresh workbook from the current database state."""
        export_path = target_path or self.workbook_path
        workbook = Workbook()
        worksheet = workbook.active

        worksheet.title = "Inventory"
        worksheet.append(EXCEL_HEADERS)

        stmt = select(VehicleStock).order_by(
            VehicleStock.branch_code, VehicleStock.model_name
        )
        result = await self.session.execute(stmt)
        for stock in result.scalars():
            worksheet.append([
                stock.branch_code,
                stock.branch_name,
                stock.city,
                self._round_float(stock.latitude),
                self._round_float(stock.longitude),
                stock.model_code,
                stock.model_name,
                stock.variant,
                stock.color,
                int(stock.quantity),
                int(stock.reserved),
            ])

        self._autosize_columns(worksheet)
        worksheet.freeze_panes = "A2"

        await asyncio.to_thread(workbook.save, export_path)
        await asyncio.to_thread(workbook.close)
        return export_path

    async def _get_stock_by_row(self, row_number: int) -> VehicleStock | None:
        stmt = select(VehicleStock).where(VehicleStock.excel_row_number == row_number)
        result = await self.session.execute(stmt)
        return result.scalar_one_or_none()

    async def _ensure_branch(self, payload: dict) -> Branch | None:
        code = payload.get("branch_code")
        name = payload.get("branch_name")
        city = payload.get("city")

        if not code or not name or not city:
            return None

        stmt = select(Branch).where(Branch.code == code)
        result = await self.session.execute(stmt)
        branch = result.scalar_one_or_none()
        if branch is None:
            branch = Branch(
                code=code,
                name=name,
                city=city,
                latitude=payload.get("latitude"),
                longitude=payload.get("longitude"),
            )
            self.session.add(branch)
            await self.session.flush()
        else:
            updated = False
            if payload.get("latitude") is not None and branch.latitude != payload["latitude"]:
                branch.latitude = payload["latitude"]
                updated = True
            if payload.get("longitude") is not None and branch.longitude != payload["longitude"]:
                branch.longitude = payload["longitude"]
                updated = True
            if name and branch.name != name:
                branch.name = name
                updated = True
            if city and branch.city != city:
                branch.city = city
                updated = True
            if updated:
                await self.session.flush()
        return branch

    def _payload_from_row(self, row: dict, excel_row_number: int) -> dict:
        return {
            "excel_row_number": excel_row_number,
            "branch_code": self._safe_str(row.get("branch_code")),
            "branch_name": self._safe_str(row.get("branch_name")),
            "city": self._safe_str(row.get("city")),
            "latitude": self._safe_float(row.get("latitude")),
            "longitude": self._safe_float(row.get("longitude")),
            "model_code": self._safe_str(row.get("model_code")),
            "model_name": self._safe_str(row.get("model_name")) or "Unknown Model",
            "variant": self._safe_str(row.get("variant")),
            "color": self._safe_str(row.get("color")),
            "quantity": self._safe_int(row.get("quantity"), default=0),
            "reserved": self._safe_int(row.get("reserved"), default=0),
            "last_synced_at": datetime.utcnow(),
        }

    async def _load_workbook(self) -> Workbook:
        if not self.workbook_path.exists():
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Inventory"
            sheet.append(EXCEL_HEADERS)
            await asyncio.to_thread(workbook.save, self.workbook_path)
            await asyncio.to_thread(workbook.close)
        return await asyncio.to_thread(load_workbook, self.workbook_path)

    @staticmethod
    def _autosize_columns(worksheet) -> None:
        for idx, column in enumerate(worksheet.columns, start=1):
            max_length = 0
            column_letter = get_column_letter(idx)
            for cell in column:
                try:
                    value = str(cell.value) if cell.value is not None else ""
                    max_length = max(max_length, len(value))
                except Exception:
                    continue
            worksheet.column_dimensions[column_letter].width = max(max_length + 2, 12)

    @staticmethod
    def _normalize(value) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip().lower().replace(" ", "_")
        return str(value).strip().lower()

    @staticmethod
    def _safe_str(value) -> str | None:
        if value in (None, ""):
            return None
        return str(value).strip()

    @staticmethod
    def _safe_float(value) -> float | None:
        if value in (None, ""):
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _safe_int(value, default: int = 0) -> int:
        try:
            if value in (None, ""):
                return default
            return int(float(value))
        except (TypeError, ValueError):
            return default

    @staticmethod
    def _round_float(value) -> float | None:
        if value is None:
            return None
        return round(float(value), 6)
