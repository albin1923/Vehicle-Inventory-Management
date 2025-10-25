from __future__ import annotations

import csv
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.branch import Branch
from app.models.import_log import ImportJob
from app.models.inventory import Inventory
from app.models.vehicle_model import VehicleModel
from app.schemas.import_job import ImportJobCreate


class ImportService:
    def __init__(self, session: AsyncSession) -> None:
        self.session = session

    async def queue_import(
        self,
        payload: ImportJobCreate,
        file_buffer: BytesIO,
        uploaded_by_id: int | None = None,
    ) -> ImportJob:
        storage_path = self._persist_file(payload.source_filename, file_buffer)

        rows = self._load_rows(storage_path, payload.sheet_name)
        if not rows:
            raise ValueError("Uploaded file does not contain any data rows")

        summary = await self._apply_inventory_updates(rows)

        job_status = "completed" if not summary["errors"] else "completed_with_issues"
        job = ImportJob(
            branch_id=payload.branch_id,
            uploaded_by_id=uploaded_by_id,
            source_filename=payload.source_filename,
            sheet_name=payload.sheet_name,
            status=job_status,
            summary={
                "stored_path": storage_path.as_posix(),
                "processed_rows": summary["processed_rows"],
                "total_rows": summary["processed_rows"],
                "updated_inventory": summary["updated"],
                "created_inventory": summary["created"],
                "branches_created": summary["branches_created"],
                "branches_updated": summary["branches_updated"],
                "error_count": len(summary["errors"]),
                "errors": summary["errors"],
            },
            executed_at=datetime.now(timezone.utc),
        )
        self.session.add(job)
        await self.session.commit()
        await self.session.refresh(job)
        return job

    async def list_recent(self, limit: int = 25) -> list[ImportJob]:
        stmt = select(ImportJob).order_by(ImportJob.created_at.desc()).limit(limit)
        result = await self.session.execute(stmt)
        return list(result.scalars().all())

    @staticmethod
    def _persist_file(filename: str, buffer: BytesIO) -> Path:
        base_dir = Path("storage") / "imports"
        base_dir.mkdir(parents=True, exist_ok=True)

        safe_name = filename.replace("/", "_").replace("\\", "_")
        target_path = base_dir / safe_name
        counter = 1
        while target_path.exists():
            stem = Path(safe_name).stem
            suffix = Path(safe_name).suffix
            target_path = base_dir / f"{stem}_{counter}{suffix}"
            counter += 1

        with target_path.open("wb") as f:
            f.write(buffer.getvalue())

        return target_path

    def _load_rows(self, file_path: Path, sheet_name: str | None) -> list[dict[str, Any]]:
        suffix = file_path.suffix.lower()
        if suffix == ".csv":
            return self._load_csv_rows(file_path)

        if suffix in {".xlsx", ".xlsm"}:
            return self._load_excel_rows(file_path, sheet_name)

        raise ValueError(f"Unsupported file type '{suffix}'. Upload .csv or .xlsx files")

    def _load_csv_rows(self, file_path: Path) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            if reader.fieldnames is None:
                return rows
            normalized_fieldnames = [self._normalize_header(name) for name in reader.fieldnames if name]
            for raw_row in reader:
                normalized_row = {
                    self._normalize_header(key): value.strip() if isinstance(value, str) else value
                    for key, value in raw_row.items()
                }
                rows.append({key: normalized_row.get(key) for key in normalized_fieldnames})
        return rows

    def _load_excel_rows(self, file_path: Path, sheet_name: str | None) -> list[dict[str, Any]]:
        workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
        worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active

        iterator = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(iterator)
        except StopIteration:
            return []

        normalized_header = [self._normalize_header(cell) for cell in header_row]
        rows: list[dict[str, Any]] = []
        for row in iterator:
            row_dict = {normalized_header[idx]: row[idx] for idx in range(len(normalized_header)) if normalized_header[idx]}
            cleaned = {
                key: (value.strip() if isinstance(value, str) else value)
                for key, value in row_dict.items()
            }
            if any(value not in (None, "") for value in cleaned.values()):
                rows.append(cleaned)
        workbook.close()
        return rows

    async def _apply_inventory_updates(self, rows: Iterable[dict[str, Any]]) -> dict[str, Any]:
        branch_map = await self._load_branch_map()
        model_code_map, model_name_map = await self._load_model_maps()

        processed = 0
        created = 0
        updated = 0
        branches_created = 0
        branches_updated = 0
        created_branch_codes: set[str] = set()
        updated_branch_codes: set[str] = set()
        missing_coordinates: set[str] = set()
        errors: list[dict[str, Any]] = []

        for idx, raw_row in enumerate(rows, start=1):
            processed += 1
            try:
                branch_code = self._require_field(raw_row, "branch_code")
                model_code = self._require_field(raw_row, "model_code")
            except ValueError as exc:
                errors.append({"row": idx, "detail": str(exc)})
                continue

            branch = branch_map.get(branch_code)
            if branch is None:
                branch_name = self._optional_field(raw_row, "branch_name")
                branch_city = self._optional_field(raw_row, "city")
                if branch_name and branch_city:
                    latitude = self._optional_float(raw_row, "latitude")
                    longitude = self._optional_float(raw_row, "longitude")
                    
                    # Validate coordinate ranges
                    if latitude is not None and not (-90 <= latitude <= 90):
                        errors.append({"row": idx, "detail": f"Latitude {latitude} out of range (-90 to 90)"})
                        continue
                    if longitude is not None and not (-180 <= longitude <= 180):
                        errors.append({"row": idx, "detail": f"Longitude {longitude} out of range (-180 to 180)"})
                        continue
                    
                    branch = Branch(
                        code=branch_code,
                        name=branch_name,
                        city=branch_city,
                        latitude=latitude,
                        longitude=longitude,
                    )
                    self.session.add(branch)
                    await self.session.flush()
                    branch_map[branch_code] = branch
                    if branch_code not in created_branch_codes:
                        created_branch_codes.add(branch_code)
                        branches_created += 1
                else:
                    errors.append({
                        "row": idx,
                        "detail": (
                            f"Unknown branch code '{branch_code}'. Provide branch_name and city (plus latitude/longitude for nearest calculation)."
                        ),
                    })
                    continue

            else:
                latitude = self._optional_float(raw_row, "latitude")
                longitude = self._optional_float(raw_row, "longitude")
                
                # Validate coordinate ranges when updating
                if latitude is not None and not (-90 <= latitude <= 90):
                    errors.append({"row": idx, "detail": f"Latitude {latitude} out of range (-90 to 90)"})
                    continue
                if longitude is not None and not (-180 <= longitude <= 180):
                    errors.append({"row": idx, "detail": f"Longitude {longitude} out of range (-180 to 180)"})
                    continue
                
                has_coordinate_update = False
                if latitude is not None and branch.latitude != latitude:
                    branch.latitude = latitude
                    has_coordinate_update = True
                if longitude is not None and branch.longitude != longitude:
                    branch.longitude = longitude
                    has_coordinate_update = True
                branch_name = self._optional_field(raw_row, "branch_name")
                branch_city = self._optional_field(raw_row, "city")
                if branch_name and branch.name != branch_name:
                    branch.name = branch_name
                    has_coordinate_update = True
                if branch_city and branch.city != branch_city:
                    branch.city = branch_city
                    has_coordinate_update = True
                if has_coordinate_update and branch_code not in updated_branch_codes:
                    updated_branch_codes.add(branch_code)
                    branches_updated += 1

            if (branch.latitude is None or branch.longitude is None) and branch_code not in missing_coordinates:
                errors.append({
                    "row": idx,
                    "detail": (
                        f"Branch '{branch_code}' is missing latitude/longitude. Nearest showroom calculations require coordinates."
                    ),
                })
                missing_coordinates.add(branch_code)

            quantity_raw = raw_row.get("quantity")
            reserved_raw = raw_row.get("reserved", 0)

            try:
                quantity = int(float(quantity_raw))
                reserved = int(float(reserved_raw or 0))
            except (TypeError, ValueError):
                errors.append({"row": idx, "detail": "Quantity or reserved value is not a number"})
                continue

            if quantity < 0 or reserved < 0:
                errors.append({"row": idx, "detail": "Quantity and reserved must be non-negative"})
                continue

            model = model_code_map.get(model_code)
            if model is None:
                model_name = raw_row.get("model_name") or raw_row.get("model") or model_code
                normalized_model_name = str(model_name).strip().lower()
                model = model_name_map.get(normalized_model_name)

                if model is None:
                    model = VehicleModel(name=str(model_name), external_code=model_code)
                    self.session.add(model)
                    await self.session.flush()
                elif model.external_code is None:
                    # Attach the newly seen external code for existing models missing a code.
                    model.external_code = model_code

                model_code_map[model_code] = model
                model_name_map[normalized_model_name] = model

            inventory = await self._fetch_inventory(branch.id, model.id)
            if inventory is None:
                inventory = Inventory(branch_id=branch.id, model_id=model.id, quantity=quantity, reserved=reserved)
                self.session.add(inventory)
                created += 1
            else:
                inventory.quantity = quantity
                inventory.reserved = reserved
                updated += 1

        await self.session.commit()

        return {
            "processed_rows": processed,
            "created": created,
            "updated": updated,
            "branches_created": branches_created,
            "branches_updated": branches_updated,
            "errors": errors,
        }

    async def _load_branch_map(self) -> dict[str, Branch]:
        result = await self.session.execute(select(Branch))
        return {branch.code: branch for branch in result.scalars().all()}

    async def _load_model_maps(self) -> tuple[dict[str, VehicleModel], dict[str, VehicleModel]]:
        result = await self.session.execute(select(VehicleModel))
        code_map: dict[str, VehicleModel] = {}
        name_map: dict[str, VehicleModel] = {}
        for model in result.scalars().all():
            if model.external_code:
                code_map[model.external_code] = model
            name_map[model.name.strip().lower()] = model
        return code_map, name_map

    async def _fetch_inventory(self, branch_id: int, model_id: int) -> Inventory | None:
        stmt = select(Inventory).where(Inventory.branch_id == branch_id, Inventory.model_id == model_id)
        result = await self.session.execute(stmt)
        return result.scalar_one_or_none()

    @staticmethod
    def _normalize_header(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            normalized = value.strip().lower().replace(" ", "_").replace("-", "_")
            return normalized
        return str(value).strip().lower()

    @staticmethod
    def _require_field(row: dict[str, Any], key: str) -> str:
        value = row.get(key)
        if value in (None, ""):
            raise ValueError(f"Missing required column '{key}'")
        return str(value).strip()

    @staticmethod
    def _optional_field(row: dict[str, Any], key: str) -> str | None:
        value = row.get(key)
        if value in (None, ""):
            return None
        return str(value).strip()

    @staticmethod
    def _optional_float(row: dict[str, Any], key: str) -> float | None:
        value = row.get(key)
        if value in (None, ""):
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
